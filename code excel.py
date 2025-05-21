import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict

# Ruta al archivo Excel (ajusta la ruta a la correcta)
file_path = 'C://Users//Bruno//Desktop//CHACO INTERNACIONAL - DISTRIBUCION.xlsx'
output_file = 'resultado_analisis.txt'  # Nombre del archivo de salida

try:
    # Abrir el archivo de salida
    with open(output_file, 'w', encoding='utf-8') as f:
        # Leer el archivo Excel con pandas
        xls = pd.ExcelFile(file_path)

        # Listar las hojas
        hojas = xls.sheet_names
        f.write("Hojas encontradas:\n")
        f.write(", ".join(hojas) + "\n\n")

        # Análisis de cada hoja
        for hoja in hojas:
            f.write(f"\nAnálisis de la hoja: {hoja}\n")
            df = pd.read_excel(xls, sheet_name=hoja)
            f.write(str(df.head()) + "\n")  # Muestra las primeras filas de la hoja

            # Cargar el libro de trabajo para análisis más profundo
            wb = load_workbook(file_path, data_only=False)
            ws = wb[hoja]

            # Diccionario para almacenar valores y sus coordenadas
            valores = defaultdict(list)

            # Análisis de celdas y fórmulas (omitimos celdas vacías)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:  # Solo procesar celdas no vacías
                        valores[cell.value].append(cell.coordinate)

            # Escribir resultados agrupando celdas con valores iguales
            for valor, celdas in valores.items():
                if len(celdas) > 1:
                    f.write(f"Las celdas {', '.join(celdas)} tienen el mismo valor: {valor}\n")
                else:
                    f.write(f"Célula {celdas[0]} tiene valor: {valor}\n")

        # Resumen de resultados
        f.write("\nAnálisis completo.\n")

    print(f"Análisis completado. Resultado guardado en '{output_file}'")

except FileNotFoundError:
    print(f"No se encontró el archivo en la ruta especificada: {file_path}")
except Exception as e:
    print(f"Ocurrió un error: {str(e)}")

